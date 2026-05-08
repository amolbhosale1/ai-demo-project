from pathlib import Path

from openpyxl import load_workbook

from stakeholder_mapper_agent.graph import _select_connector
from stakeholder_mapper_agent.output.stakeholder_markdown import write_stakeholder_markdown
from stakeholder_mapper_agent.scoring import merge_dedupe_and_score
from stakeholder_mapper_agent.output.stakeholder_xlsx import FIELDS, write_stakeholder_xlsx
from stakeholder_mapper_agent.types import OrganizationProfile, Stakeholder


def test_connector_selection_by_country_system():
    assert _select_connector("uk_parliamentary").__class__.__name__ == "UKConnector"
    assert _select_connector("india_parliamentary").__class__.__name__ == "IndiaConnector"
    assert _select_connector("something_else").__class__.__name__ == "GenericConnector"


def test_india_connector_uses_state_scope_region_label():
    profile = OrganizationProfile(
        client_name="Client",
        country="India",
        country_system="india_parliamentary",
        geography_scope="state",
        target_regions=["Maharashtra", "Gujarat"],
        services=[],
        target_sectors=[],
        priority_topics=[],
        geo_focus=["India"],
    )
    connector = _select_connector("india_parliamentary")
    rows = connector.find_executive(profile, ["AI governance"])
    assert rows
    assert rows[0].constituency_or_region == "Maharashtra, Gujarat"


def test_merge_dedupe_and_score_keeps_highest_score():
    profile = OrganizationProfile(
        client_name="Client",
        country="India",
        country_system="india_parliamentary",
        geography_scope="state",
        target_regions=["Maharashtra"],
        services=["MVP development", "AI agent development"],
        target_sectors=["B2B SaaS"],
        priority_topics=["AI governance", "startup policy"],
        geo_focus=["India"],
    )
    a = Stakeholder(
        name="Example Committee",
        stakeholder_type="Committee",
        constituency_or_region="India",
        country="India",
        contact_information="contact",
        policy_area="AI governance",
        reason_for_relevance="Covers AI policy oversight.",
        evidence_urls=["https://example.org/a"],
        source_bucket="Groups_and_Committees",
    )
    b = Stakeholder(
        name="Example Committee",
        stakeholder_type="Committee",
        constituency_or_region="India",
        country="India",
        contact_information="contact",
        policy_area="AI governance",
        reason_for_relevance="Covers startup policy and AI governance.",
        evidence_urls=["https://example.org/b"],
        source_bucket="Groups_and_Committees",
    )
    merged = merge_dedupe_and_score([a, b], profile)
    assert len(merged) == 1
    assert merged[0].relevance_score >= 1.0


def test_stakeholder_xlsx_has_required_headers(tmp_path: Path):
    rows = [
        Stakeholder(
            name="A",
            stakeholder_type="Minister",
            constituency_or_region="UK",
            country="United Kingdom",
            contact_information="https://example.org",
            policy_area="public affairs",
            reason_for_relevance="Relevant to policy area.",
            evidence_urls=["https://example.org/evidence"],
            source_bucket="Executive_Stakeholders",
            relevance_score=2.0,
        )
    ]
    out = write_stakeholder_xlsx(tmp_path, "Client", rows)
    wb = load_workbook(out)
    ws = wb["Executive_Stakeholders"]
    headers = [c.value for c in ws[1]]
    assert headers == FIELDS
    assert "Audit_Log" in wb.sheetnames


def test_stakeholder_markdown_contains_mermaid_and_details(tmp_path: Path):
    rows = [
        Stakeholder(
            name="Example Stakeholder",
            stakeholder_type="Minister",
            constituency_or_region="London",
            country="United Kingdom",
            contact_information="https://example.org",
            policy_area="public policy",
            reason_for_relevance="Relevant to policy area.",
            evidence_urls=["https://example.org/evidence"],
            source_bucket="Executive_Stakeholders",
            relevance_score=2.0,
        )
    ]
    out = write_stakeholder_markdown(tmp_path, "Client", rows)
    content = out.read_text(encoding="utf-8")
    assert "```mermaid" in content
    assert "Relationship Network" in content
    assert "Example Stakeholder" in content

