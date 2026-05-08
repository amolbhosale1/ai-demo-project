from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from stakeholder_mapper_agent.types import Stakeholder

FIELDS = [
    "Name",
    "Stakeholder Type",
    "Constituency or equivalent",
    "Country",
    "Contact information",
    "Policy area",
    "Reason for relevance",
    "Evidence URL(s)",
    "Relevance score",
]


def write_stakeholder_xlsx(output_dir: str | Path, client_name: str, stakeholders: list[Stakeholder]) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    now_utc = datetime.now(UTC)
    filename = f"{client_name.lower().replace(' ', '_')}_stakeholders_{now_utc.date()}.xlsx"
    file_path = output_path / filename

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_bucket_sheet(wb, "Executive_Stakeholders", [s for s in stakeholders if s.source_bucket == "Executive_Stakeholders"])
    _write_bucket_sheet(wb, "Parliamentary_Stakeholders", [s for s in stakeholders if s.source_bucket == "Parliamentary_Stakeholders"])
    _write_bucket_sheet(wb, "Groups_and_Committees", [s for s in stakeholders if s.source_bucket == "Groups_and_Committees"])
    _write_audit_sheet(wb, stakeholders, now_utc.isoformat())

    wb.save(file_path)
    return file_path


def _write_bucket_sheet(wb: Workbook, title: str, rows: list[Stakeholder]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(FIELDS)
    for s in rows:
        ws.append(
            [
                s.name,
                s.stakeholder_type,
                s.constituency_or_region,
                s.country,
                s.contact_information,
                s.policy_area,
                s.reason_for_relevance,
                ", ".join(s.evidence_urls),
                s.relevance_score,
            ]
        )


def _write_audit_sheet(wb: Workbook, stakeholders: list[Stakeholder], generated_at: str) -> None:
    ws = wb.create_sheet(title="Audit_Log")
    ws.append(["Generated At (UTC)", generated_at])
    ws.append(["Total Stakeholders", len(stakeholders)])
    ws.append([])
    ws.append(["Name", "Source Bucket", "Evidence URL(s)"])
    for s in stakeholders:
        ws.append([s.name, s.source_bucket, ", ".join(s.evidence_urls)])

